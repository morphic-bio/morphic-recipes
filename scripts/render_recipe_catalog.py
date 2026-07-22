#!/usr/bin/env python3
"""Render the recipe catalog into human-readable views.

catalog.yaml is the single source of truth (agent-friendly). This regenerates the
human views RECIPES.md (GitHub-rendered table) and RECIPES.xlsx (spreadsheet) from
it, so the views never drift. Run after any catalog.yaml change:

    python3 scripts/render_recipe_catalog.py

Requires pyyaml + openpyxl (both present in the suite env).
"""
from __future__ import annotations
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
CATALOG = REPO / "catalog.yaml"
MD_OUT = REPO / "RECIPES.md"
XLSX_OUT = REPO / "RECIPES.xlsx"

COLUMNS = [
    ("id", "id"),
    ("title", "recipe"),
    ("modality", "modality"),
    ("engine", "engine"),
    ("minimal_wrapper", "minimal wrapper"),
    ("profiles", "profiles"),
    ("compose_up", "compose-up"),
    ("provenance_oracle", "provenance oracle"),
    ("composition_examples", "composition examples"),
    ("status", "status"),
]

GENERATED = "GENERATED from catalog.yaml by scripts/render_recipe_catalog.py — do not hand-edit."


def fmt(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)


def compose_badge(value) -> str:
    if value is True:
        return "✓"
    if value is False:
        return "—"
    return str(value)  # e.g. "inherits"


def composition_examples(value) -> str:
    if not isinstance(value, list):
        return ""
    return ", ".join(
        str(item.get("graph_id") or "") if isinstance(item, dict) else str(item)
        for item in value
        if (item.get("graph_id") if isinstance(item, dict) else item)
    )


def render_markdown(cat: dict) -> str:
    recipes = cat.get("recipes", [])
    L = []
    L.append("# Recipe catalog — starting points")
    L.append("")
    L.append(f"> {GENERATED}")
    L.append(">")
    L.append("> The curated single source of truth is [`catalog.yaml`](catalog.yaml).")
    L.append("> This list is deliberately **small**: it holds canonical *starting points*,")
    L.append("> not a record of every run (that is `morphic-provenance`). Provenance is the")
    L.append("> oracle for parameter *values*; a recipe's `--profile`/compose-up governs which")
    L.append("> output *layers* it emits. See AGENTS.md \"Compose to the target\".")
    L.append("")
    # compact table
    head = "| " + " | ".join(h for _, h in COLUMNS) + " |"
    sep = "|" + "|".join("---" for _ in COLUMNS) + "|"
    L.append(head)
    L.append(sep)
    for r in recipes:
        cells = []
        for key, _ in COLUMNS:
            if key == "compose_up":
                cells.append(compose_badge(r.get(key)))
            elif key == "composition_examples":
                cells.append(composition_examples(r.get(key)) or "—")
            elif key in ("engine", "minimal_wrapper"):
                v = fmt(r.get(key))
                cells.append(f"`{v}`" if v else "—")
            else:
                cells.append(fmt(r.get(key)) or "—")
        L.append("| " + " | ".join(cells) + " |")
    L.append("")
    # notes
    L.append("## Notes")
    L.append("")
    for r in recipes:
        note = " ".join(fmt(r.get("notes")).split())
        if note:
            L.append(f"- **{r['id']}** — {note}")
    L.append("")
    # excluded
    excl = cat.get("excluded", {})
    if excl:
        L.append("## Not catalogued as starting points")
        L.append("")
        L.append("Deliberately excluded to keep the list small (internal steps, smokes, "
                 "remote executors, preflight, ops):")
        L.append("")
        for kind, items in excl.items():
            L.append(f"- *{kind}*: " + ", ".join(f"`{i}`" for i in items))
        L.append("")
    # elsewhere
    elsewhere = cat.get("elsewhere", [])
    if elsewhere:
        L.append("## Elsewhere (suite-repo recipes / workflows)")
        L.append("")
        L.append("Catalogued by each MCP server's `list_workflows`; cross-reference, do not duplicate:")
        L.append("")
        for e in elsewhere:
            L.append(f"- {e}")
        L.append("")
    return "\n".join(L) + "\n"


def render_xlsx(cat: dict) -> None:
    recipes = cat.get("recipes", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Recipes"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    headers = [h for _, h in COLUMNS] + ["notes"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="top", horizontal="left")
    for r in recipes:
        row = []
        for key, _ in COLUMNS:
            if key == "compose_up":
                row.append(compose_badge(r.get(key)))
            elif key == "composition_examples":
                row.append(composition_examples(r.get(key)))
            else:
                row.append(fmt(r.get(key)))
        row.append(" ".join(fmt(r.get("notes")).split()))
        ws.append(row)
    # column widths + wrap notes
    widths = [16, 46, 12, 42, 34, 22, 11, 22, 48, 16, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    # excluded sheet
    excl = cat.get("excluded", {})
    if excl:
        ws2 = wb.create_sheet("Excluded")
        ws2.append(["category", "script"])
        for c in range(1, 3):
            ws2.cell(row=1, column=c).font = hdr_font
            ws2.cell(row=1, column=c).fill = hdr_fill
        for kind, items in excl.items():
            for it in items:
                ws2.append([kind, it])
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 48
        ws2.freeze_panes = "A2"
    wb.save(XLSX_OUT)


def main() -> int:
    cat = yaml.safe_load(CATALOG.read_text())
    MD_OUT.write_text(render_markdown(cat))
    render_xlsx(cat)
    n = len(cat.get("recipes", []))
    print(f"Rendered {n} recipes -> {MD_OUT.name}, {XLSX_OUT.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
